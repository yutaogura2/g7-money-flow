import json
import re
from pathlib import Path

import pytest

import build

SAMPLE = {
    "meta": {"last_updated": "2026-06-22", "sources": [], "disclaimer": "参考情報"},
    "meetings": [
        {
            "id": "m1", "date": "2025-05-22", "type": "財務相・中銀総裁会議",
            "presidency": "カナダ", "title_orig": "G7 FMCBG Statement",
            "title_ja": "G7財務相・中銀総裁声明", "source_url": "https://example.com/a",
            "source_lang": "en", "summary_ja": "要約A",
            "key_points_ja": ["決定1", "決定2"], "full_translation_ja": "全訳A",
            "analysis": {"fx_rates": "為替A", "commodities": "商品A",
                         "sectors": "セクターA", "geopolitics": "地政学A"},
            "market_direction": "リスクオフ", "tags": ["制裁", "為替"],
        }
    ],
    "calendar": [
        {"date": "2026-06-15", "name_ja": "G7首脳会議", "category": "G7首脳",
         "status": "予定", "related_meeting_id": None, "note": "議長国フランス"}
    ],
}


@pytest.fixture
def sample_json(tmp_path):
    p = tmp_path / "data.json"
    p.write_text(json.dumps(SAMPLE, ensure_ascii=False), encoding="utf-8")
    return p


def test_write_data_js_has_prefix_and_valid_json(tmp_path, sample_json):
    data = build.load_data(sample_json)
    out = tmp_path / "data.js"
    build.write_data_js(data, out)
    text = out.read_text(encoding="utf-8")
    assert text.startswith("window.G7DATA = ")
    body = re.sub(r"^window\.G7DATA = ", "", text).rstrip().rstrip(";")
    parsed = json.loads(body)
    assert parsed["meetings"][0]["title_ja"] == "G7財務相・中銀総裁声明"
    # 日本語がエスケープされず読めること
    assert "\\u" not in text


from openpyxl import load_workbook


def test_build_xlsx_sheets_and_rows(tmp_path, sample_json):
    data = build.load_data(sample_json)
    out = tmp_path / "out.xlsx"
    build.build_xlsx(data, out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["会議ログ", "カレンダー", "考察サマリ", "マクロ_最新", "マクロ_時系列"]
    assert wb["会議ログ"].max_row == 2       # ヘッダ + 1会議
    assert wb["カレンダー"].max_row == 2      # ヘッダ + 1イベント
    assert wb["考察サマリ"].max_row == 5      # ヘッダ + 4観点
    # ヘッダ確認
    assert wb["会議ログ"].cell(row=1, column=1).value == "日付"


def test_build_xlsx_source_hyperlink(tmp_path, sample_json):
    data = build.load_data(sample_json)
    out = tmp_path / "out.xlsx"
    build.build_xlsx(data, out)
    wb = load_workbook(out)
    ws = wb["会議ログ"]
    url_cell = ws.cell(row=2, column=len(build.MEETING_HEADERS))
    assert url_cell.hyperlink is not None
    assert url_cell.hyperlink.target == "https://example.com/a"


def _write_macro_fixture(tmp_path):
    (tmp_path / "macro_data").mkdir()
    (tmp_path / "macro_data" / "M2SL.csv").write_text(
        "date,value,yoy_pct\n2019-01-01,100.0,\n2020-01-01,110.0,10.0\n", encoding="utf-8")
    (tmp_path / "series_config.json").write_text(
        json.dumps({"history_start": "2000-01-01", "history_points": 180,
                    "series": [{"id": "M2SL", "country": "米国",
                                "indicator": "マネーサプライ(M2)", "unit": "10億ドル",
                                "transform": "yoy_pct_also"}]}), encoding="utf-8")


def test_build_macro_payload(tmp_path):
    _write_macro_fixture(tmp_path)
    payload = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    s = payload["series"][0]
    assert s["country"] == "米国"
    assert s["latest"] == 110.0
    assert s["latest_date"] == "2020-01-01"
    assert s["yoy"] == 10.0
    assert s["history"] == [["2019-01-01", 100.0], ["2020-01-01", 110.0]]
    assert s["history_yoy"] == [["2020-01-01", 10.0]]


def test_macro_timeseries_rows(tmp_path):
    _write_macro_fixture(tmp_path)
    rows = build.macro_timeseries_rows(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    assert ("2020-01-01", "米国", "マネーサプライ(M2)", 110.0, 10.0) in rows


def test_build_xlsx_includes_macro_sheets(tmp_path, sample_json, monkeypatch):
    _write_macro_fixture(tmp_path)
    monkeypatch.setattr(build, "SERIES_CONFIG", tmp_path / "series_config.json")
    monkeypatch.setattr(build, "MACRO_DIR", tmp_path / "macro_data")
    data = build.load_data(sample_json)
    data["macro"] = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    out = tmp_path / "out.xlsx"
    build.build_xlsx(data, out)
    wb = load_workbook(out)
    assert "マクロ_最新" in wb.sheetnames
    assert "マクロ_時系列" in wb.sheetnames
    assert wb["マクロ_最新"].cell(row=1, column=1).value == "国"
    assert wb["マクロ_最新"].max_row == 2          # ヘッダ + 1シリーズ
    assert wb["マクロ_時系列"].max_row == 3         # ヘッダ + 2行


def test_build_macro_payload_includes_group(tmp_path):
    _write_macro_fixture(tmp_path)
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    assert p["series"][0]["group"] == "fundamental"


def test_series_stats():
    s = build._series_stats([1.0, 2.0, 3.0, 4.0, 5.0])
    assert s["pctile"] == 100
    assert abs(s["zscore"] - 1.41) < 0.01      # (5-3)/sqrt(2)
    assert abs(s["std"] - 1.41421) < 0.001
    flat = build._series_stats([2.0, 2.0, 2.0])
    assert flat["zscore"] == 0.0


def test_value_on_or_before():
    rows = [("2026-04-01", 1.0, None), ("2026-05-01", 2.0, None), ("2026-06-01", 3.0, None)]
    assert build._value_on_or_before(rows, "2026-05-15") == 2.0
    assert build._value_on_or_before(rows, "2026-06-01") == 3.0
    assert build._value_on_or_before(rows, "2026-03-01") is None


def test_payload_zscore_stress_delta(tmp_path):
    (tmp_path / "macro_data").mkdir()
    (tmp_path / "macro_data" / "X.csv").write_text(
        "date,value\n2026-01-01,1\n2026-02-01,2\n2026-03-01,3\n2026-04-01,4\n2026-06-01,5\n",
        encoding="utf-8")
    (tmp_path / "series_config.json").write_text(json.dumps({
        "history_points": 180, "series": [
            {"id": "X", "country": "c", "indicator": "i", "unit": "u",
             "transform": "level", "risk_dir": "high"}]}), encoding="utf-8")
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    s = p["series"][0]
    assert s["pctile"] == 100
    assert abs(s["zscore"] - 1.41) < 0.01
    assert abs(s["stress"] - 1.41) < 0.01            # high → +z
    assert s["delta"] == 1.0                          # 06-01の約30日前=04-01の4 → 5-4
    assert s["delta_z"] != 0.0


def test_build_signals_regime_and_movers():
    series = [
        {"id": "A", "country": "a", "indicator": "i", "latest": 5, "stress": 1.0, "delta": 1.0, "delta_z": 2.0},
        {"id": "B", "country": "b", "indicator": "i", "latest": 3, "stress": 0.5, "delta": -0.2, "delta_z": -0.3},
        {"id": "C", "country": "c", "indicator": "i", "latest": 2, "stress": None, "delta": 0.1, "delta_z": 0.1},
    ]
    sig = build.build_signals(series)
    assert sig["regime"]["score"] == 0.75          # mean(1.0, 0.5)
    assert sig["regime"]["label"] == "リスクオフ"
    assert sig["regime"]["level"] == "warn"
    assert sig["movers"][0]["id"] == "A"           # |delta_z|=2 が最大
    assert sig["movers"][0]["dir"] == "up"


def test_payload_includes_signals(tmp_path):
    _write_macro_fixture(tmp_path)
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    assert "signals" in p
    assert "regime" in p["signals"] and "movers" in p["signals"]


from datetime import date as _date


def test_staleness_monthly_and_quarterly():
    monthly = [(f"2026-{m:02d}-01", 1.0, None) for m in range(1, 6)]  # Jan..May
    assert build._staleness(monthly, _date(2026, 6, 5))[1] is False   # ~35日
    assert build._staleness(monthly, _date(2026, 10, 1))[1] is True   # ~153日
    quarterly = [("2025-09-01", 1.0, None), ("2025-12-01", 1.0, None), ("2026-03-01", 1.0, None)]
    assert build._staleness(quarterly, _date(2026, 6, 5))[1] is False  # ~96日 < 91*2.5+7
    assert build._staleness([("2026-01-01", 1.0, None)], _date(2026, 6, 5)) == (0, False)


def test_build_briefing_lines():
    series = [
        {"indicator": "ボラティリティ", "country": "VIX", "zscore": 2.5, "stale": False, "stale_days": 3},
        {"indicator": "建玉(COT)", "country": "円", "zscore": -2.18, "stale": False, "stale_days": 9},
        {"indicator": "CPI", "country": "米国", "zscore": 0.3, "stale": True, "stale_days": 400},
    ]
    signals = {"regime": {"label": "リスクオフ", "score": 0.6, "level": "warn"},
               "movers": [{"indicator": "商品", "country": "WTI原油", "delta_z": -1.2, "dir": "down"}]}
    b = build.build_briefing(series, signals, "2026-06-25")
    text = " ".join(b["lines"])
    assert b["as_of"] == "2026-06-25"
    assert "リスクオフ" in text
    assert "極端な水準" in text and "VIX" in text
    assert "ポジション偏り" in text and "円" in text
    assert "⚠" in text and "CPI" in text


def test_write_briefing_md(tmp_path):
    p = tmp_path / "b.md"
    build.write_briefing_md({"as_of": "2026-06-25", "lines": ["a", "b"]}, path=p)
    t = p.read_text(encoding="utf-8")
    assert "週次ブリーフィング 2026-06-25" in t and "- a" in t


def test_payload_includes_briefing(tmp_path):
    _write_macro_fixture(tmp_path)
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    assert "briefing" in p and "lines" in p["briefing"]


def test_no_stale_flag_suppresses_staleness(tmp_path):
    (tmp_path / "macro_data").mkdir()
    (tmp_path / "macro_data" / "OLD.csv").write_text(
        "date,value\n2020-01-01,1\n2020-02-01,2\n", encoding="utf-8")
    (tmp_path / "series_config.json").write_text(json.dumps({
        "history_points": 180, "series": [
            {"id": "OLD", "country": "c", "indicator": "i", "unit": "u",
             "transform": "level", "no_stale": True}]}), encoding="utf-8")
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    s = p["series"][0]
    assert s["stale"] is False and s["stale_days"] == 0


def test_build_policy_cycle_directions():
    def mk(country, latest, prior):
        return {"country": country, "indicator": "政策金利", "latest": latest,
                "latest_date": "2026-06-01",
                "history": [["2025-11-01", prior], ["2026-06-01", latest]]}
    series = [
        mk("米国", 3.63, 4.33),
        mk("日本", 0.75, 0.5),
        mk("ユーロ圏", 2.25, 2.25),
        {"country": "英国", "indicator": "政策金利", "latest": 4.0,
         "latest_date": "2026-06-01", "history": [["2026-06-01", 4.0]]},
        {"country": "x", "indicator": "CPI", "latest": 300.0,
         "latest_date": "2026-06-01", "history": [["2020-01-01", 1.0], ["2026-06-01", 300.0]]},
    ]
    pc = build.build_policy_cycle(series)
    d = {p["country"]: p for p in pc}
    assert d["米国"]["dir"] == "down" and d["米国"]["delta6m"] == -0.7
    assert d["日本"]["dir"] == "up"
    assert d["ユーロ圏"]["dir"] == "flat"
    assert "英国" not in d      # 6か月前なし → skip
    assert "x" not in d         # 政策金利以外は対象外


def test_payload_signals_policy_cycle(tmp_path):
    _write_macro_fixture(tmp_path)
    p = build.build_macro_payload(
        config_path=tmp_path / "series_config.json", data_dir=tmp_path / "macro_data")
    assert "policy_cycle" in p["signals"]


def test_build_briefing_upcoming_events():
    signals = {"regime": {"label": "中立", "score": 0.0, "level": "ok"}, "movers": []}
    events = [
        {"date": "2026-06-30", "name_ja": "FOMC", "category": "FOMC"},
        {"date": "2026-07-05", "name_ja": "日銀会合", "category": "日銀"},
        {"date": "2026-08-01", "name_ja": "遠い予定", "category": "G7財務"},
        {"date": "2026-06-20", "name_ja": "過去", "category": "G7首脳"},
    ]
    b = build.build_briefing([], signals, "2026-06-26", events=events)
    text = " ".join(b["lines"])
    assert "今後の政策・会議予定" in text
    assert "06/30 FOMC(FOMC)" in text and "07/05 日銀会合(日銀)" in text
    assert "遠い予定" not in text and "過去" not in text
    # 予定行は鮮度行より前
    idx_ev = next(i for i, l in enumerate(b["lines"]) if "今後の政策・会議予定" in l)
    idx_st = next(i for i, l in enumerate(b["lines"]) if "データ鮮度" in l)
    assert idx_ev < idx_st
    # events無しなら行なし
    b2 = build.build_briefing([], signals, "2026-06-26")
    assert all("今後の政策・会議予定" not in l for l in b2["lines"])
