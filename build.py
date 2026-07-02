#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""data.json から data.js（ブラウザ用）と Excel を生成する。"""
import csv
import json
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA_JSON = ROOT / "data.json"
DATA_JS = ROOT / "data.js"
XLSX = ROOT / "G7_一次情報ログ.xlsx"
SERIES_CONFIG = ROOT / "series_config.json"
MACRO_DIR = ROOT / "macro_data"


def load_data(path: Path = DATA_JSON) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_data_js(data: dict, path: Path = DATA_JS) -> None:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    Path(path).write_text(f"window.G7DATA = {payload};\n", encoding="utf-8")


def _read_series_csv(path: Path) -> list:
    rows = []
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        has_yoy = header is not None and "yoy_pct" in header
        for line in reader:
            if len(line) < 2 or line[1] == "":
                continue
            d = line[0]
            v = float(line[1])
            y = float(line[2]) if (has_yoy and len(line) > 2 and line[2] != "") else None
            rows.append((d, v, y))
    return rows


def _series_stats(values: list) -> dict:
    if not values:
        return {"mean": 0.0, "std": 0.0, "zscore": 0.0, "pctile": 0}
    latest = values[-1]
    n = len(values)
    mean = sum(values) / n
    std = (sum((v - mean) ** 2 for v in values) / n) ** 0.5
    z = round((latest - mean) / std, 2) if std else 0.0
    pct = round(sum(1 for v in values if v <= latest) / n * 100)
    return {"mean": mean, "std": std, "zscore": z, "pctile": pct}


def _value_on_or_before(rows: list, target: str):
    val = None
    for r in rows:
        if r[0] <= target:
            val = r[1]
        else:
            break
    return val


def _staleness(rows: list, today: date) -> tuple:
    dates = [r[0] for r in rows]
    if len(dates) < 2:
        return (0, False)
    ds = [date.fromisoformat(x) for x in dates[-13:]]
    gaps = sorted(g for g in ((ds[i] - ds[i - 1]).days for i in range(1, len(ds))) if g > 0)
    med = gaps[len(gaps) // 2] if gaps else 30
    stale_days = (today - date.fromisoformat(dates[-1])).days
    return (stale_days, stale_days > med * 2.5 + 7)


def build_signals(series: list) -> dict:
    stresses = [s["stress"] for s in series if s.get("stress") is not None]
    if stresses:
        score = round(sum(stresses) / len(stresses), 2)
        if score >= 0.5:
            label, level = "リスクオフ", "warn"
        elif score <= -0.5:
            label, level = "リスクオン", "calm"
        else:
            label, level = "中立", "ok"
    else:
        score, label, level = None, "—", "ok"
    ranked = sorted(series, key=lambda s: abs(s.get("delta_z") or 0), reverse=True)[:6]
    movers = [{
        "id": m["id"], "country": m["country"], "indicator": m["indicator"],
        "latest": m["latest"], "delta": m.get("delta", 0), "delta_z": m.get("delta_z", 0),
        "dir": "up" if (m.get("delta") or 0) >= 0 else "down",
    } for m in ranked]
    return {"regime": {"score": score, "label": label, "level": level}, "movers": movers}


def build_briefing(series: list, signals: dict, as_of: str, events=None) -> dict:
    lines = []
    reg = signals.get("regime", {})
    lines.append(f"リスクレジーム: {reg.get('label', '—')} (総合スコア {reg.get('score')})")
    movers = signals.get("movers", [])[:5]
    if movers:
        parts = [f"{'▲' if m['dir'] == 'up' else '▼'}{m['indicator']}・{m['country']}(Δz {m['delta_z']})" for m in movers]
        lines.append("直近の主な変化: " + " / ".join(parts))
    ext = sorted([s for s in series if s.get("zscore") is not None and abs(s["zscore"]) >= 2],
                 key=lambda s: abs(s["zscore"]), reverse=True)
    if ext:
        parts = [f"{s['indicator']}・{s['country']} z{s['zscore']}({'高' if s['zscore'] > 0 else '低'})" for s in ext]
        lines.append("極端な水準(|Z|≥2): " + " / ".join(parts))
    pos = sorted([s for s in series if s.get("indicator") == "建玉(COT)" and s.get("zscore") is not None and abs(s["zscore"]) >= 1.5],
                 key=lambda s: abs(s["zscore"]), reverse=True)
    if pos:
        parts = [f"{s['country']} z{s['zscore']}({'買い過熱' if s['zscore'] > 0 else '売り過熱'})" for s in pos]
        lines.append("ポジション偏り(建玉|Z|≥1.5): " + " / ".join(parts))
    if events:
        limit = (date.fromisoformat(as_of) + timedelta(days=14)).isoformat()
        upcoming = sorted([e for e in events
                           if as_of <= (e.get("date") or "") <= limit],
                          key=lambda e: e["date"])[:6]
        if upcoming:
            parts = [f"{e['date'][5:7]}/{e['date'][8:10]} {e.get('name_ja','')}({e.get('category','')})"
                     for e in upcoming]
            lines.append("今後の政策・会議予定: " + " / ".join(parts))
    stale = [s for s in series if s.get("stale")]
    if stale:
        lines.append("データ鮮度: ⚠ " + " / ".join(f"{s['indicator']}・{s['country']}({s['stale_days']}日)" for s in stale))
    else:
        lines.append("データ鮮度: 異常なし")
    return {"as_of": as_of, "lines": lines}


def write_briefing_md(briefing: dict, path: Path = ROOT / "週次ブリーフィング.md") -> None:
    md = f"# 週次ブリーフィング {briefing['as_of']}\n\n" + "\n".join(f"- {ln}" for ln in briefing["lines"]) + "\n"
    Path(path).write_text(md, encoding="utf-8")


def build_macro_payload(config_path=None, data_dir=None, points=None, calendar=None) -> dict:
    config_path = config_path or SERIES_CONFIG      # 呼び出し時に解決（monkeypatch対応）
    data_dir = data_dir or MACRO_DIR
    if not Path(config_path).exists():
        return {"series": []}
    cfg = json.loads(Path(config_path).read_text(encoding="utf-8"))
    pts = points or cfg.get("history_points", 180)
    series = []
    for s in cfg.get("series", []):
        fp = Path(data_dir) / f"{s['id']}.csv"
        if not fp.exists():
            continue
        rows = _read_series_csv(fp)
        if not rows:
            continue
        d, v, y = rows[-1]
        values = [r[1] for r in rows]
        stats = _series_stats(values)
        zscore = stats["zscore"]
        std = stats["std"]
        risk_dir = s.get("risk_dir")
        stress = (zscore if risk_dir == "high" else -zscore) if risk_dir else None
        yy, mm, dd = (int(x) for x in d.split("-"))
        target = (date(yy, mm, dd) - timedelta(days=30)).isoformat()
        prev = _value_on_or_before(rows, target)
        delta = round(v - prev, 4) if prev is not None else 0.0
        delta_z = round((v - prev) / std, 2) if (prev is not None and std) else 0.0
        stale_days, stale = _staleness(rows, date.today())
        if s.get("no_stale"):
            stale_days, stale = 0, False
        series.append({
            "id": s["id"], "country": s["country"], "indicator": s["indicator"],
            "unit": s.get("unit", ""),
            "group": s.get("group", "fundamental"),
            "latest": v, "latest_date": d, "yoy": y,
            "zscore": zscore, "pctile": stats["pctile"], "stress": stress,
            "delta": delta, "delta_z": delta_z,
            "stale_days": stale_days, "stale": stale,
            "history": [[r[0], r[1]] for r in rows][-pts:],
            "history_yoy": [[r[0], r[2]] for r in rows if r[2] is not None][-pts:],
        })
    signals = build_signals(series)
    briefing = build_briefing(series, signals, date.today().isoformat(), events=calendar)
    return {"series": series, "signals": signals, "briefing": briefing}


def macro_timeseries_rows(config_path=None, data_dir=None) -> list:
    config_path = config_path or SERIES_CONFIG      # 呼び出し時に解決（monkeypatch対応）
    data_dir = data_dir or MACRO_DIR
    if not Path(config_path).exists():
        return []
    cfg = json.loads(Path(config_path).read_text(encoding="utf-8"))
    start = cfg.get("history_start", "2000-01-01")
    out = []
    for s in cfg.get("series", []):
        fp = Path(data_dir) / f"{s['id']}.csv"
        if not fp.exists():
            continue
        for d, v, y in _read_series_csv(fp):
            if d >= start:
                out.append((d, s["country"], s["indicator"], v, y))
    out.sort(key=lambda r: (r[0], r[1], r[2]))
    return out


MEETING_HEADERS = [
    "日付", "種別", "議長国", "邦題", "原題", "要約", "決定事項",
    "主要部分の全訳", "考察(為替・金利)", "考察(商品・エネルギー・金)",
    "考察(セクター)", "考察(地政学)", "方向性メモ", "タグ", "原文URL",
]
CALENDAR_HEADERS = ["日付", "名称", "カテゴリ", "状態", "関連会議ID", "備考"]
ANALYSIS_HEADERS = ["日付", "観点", "論点", "関連会議ID"]

ASPECTS = [
    ("fx_rates", "為替・金利"),
    ("commodities", "商品・エネルギー・金"),
    ("sectors", "セクター"),
    ("geopolitics", "地政学"),
]


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_xlsx(data: dict, path: Path = XLSX) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "会議ログ"
    ws1.append(MEETING_HEADERS)
    for m in data.get("meetings", []):
        a = m.get("analysis", {})
        ws1.append([
            m.get("date", ""), m.get("type", ""), m.get("presidency", ""),
            m.get("title_ja", ""), m.get("title_orig", ""), m.get("summary_ja", ""),
            "\n".join(m.get("key_points_ja", [])), m.get("full_translation_ja", ""),
            a.get("fx_rates", ""), a.get("commodities", ""),
            a.get("sectors", ""), a.get("geopolitics", ""),
            m.get("market_direction", ""), ", ".join(m.get("tags", [])),
            m.get("source_url", ""),
        ])
        url = m.get("source_url", "")
        if url:
            c = ws1.cell(row=ws1.max_row, column=len(MEETING_HEADERS))
            c.hyperlink = url
            c.font = Font(color="0563C1", underline="single")
    _style_header(ws1)

    ws2 = wb.create_sheet("カレンダー")
    ws2.append(CALENDAR_HEADERS)
    for e in sorted(data.get("calendar", []), key=lambda x: x.get("date", "")):
        ws2.append([
            e.get("date", ""), e.get("name_ja", ""), e.get("category", ""),
            e.get("status", ""), e.get("related_meeting_id") or "", e.get("note", ""),
        ])
    _style_header(ws2)

    ws3 = wb.create_sheet("考察サマリ")
    ws3.append(ANALYSIS_HEADERS)
    for m in data.get("meetings", []):
        a = m.get("analysis", {})
        for key, label in ASPECTS:
            text = a.get(key, "")
            if text:
                ws3.append([m.get("date", ""), label, text, m.get("id", "")])
    _style_header(ws3)

    macro = data.get("macro", {}).get("series", [])
    ws4 = wb.create_sheet("マクロ_最新")
    ws4.append(["国", "指標", "最新値", "日付", "前年比%", "単位"])
    for s in macro:
        ws4.append([s["country"], s["indicator"], s["latest"], s["latest_date"],
                    "" if s.get("yoy") is None else s["yoy"], s.get("unit", "")])
    _style_header(ws4)

    ws5 = wb.create_sheet("マクロ_時系列")
    ws5.append(["日付", "国", "指標", "値", "前年比%"])
    for d, country, ind, v, y in macro_timeseries_rows():
        ws5.append([d, country, ind, v, "" if y is None else y])
    _style_header(ws5)

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    data = load_data()
    data["macro"] = build_macro_payload(calendar=data.get("calendar"))
    write_briefing_md(data["macro"]["briefing"])
    write_data_js(data)
    build_xlsx(data)
    print(f"OK: {len(data.get('meetings', []))} meetings, "
          f"{len(data.get('calendar', []))} calendar events")
    return 0


if __name__ == "__main__":
    sys.exit(main())
